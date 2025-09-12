#!/usr/bin/env python3
"""
Comprehensive Backend Testing for Family Quiz Application
Tests FastAPI server, WebSocket connections, Excel processing, QR code generation, and quiz management
"""

import asyncio
import json
import requests
import socketio
import openpyxl
from io import BytesIO
import base64
import time
from pathlib import Path
import os
from dotenv import load_dotenv

# Load environment variables
load_dotenv('/app/frontend/.env')
BACKEND_URL = os.getenv('REACT_APP_BACKEND_URL', 'https://flashquiz-2.preview.emergentagent.com')
API_BASE = f"{BACKEND_URL}/api"

class BackendTester:
    def __init__(self):
        self.session = requests.Session()
        self.test_results = {}
        
    def log_test(self, test_name, success, message=""):
        """Log test results"""
        status = "✅ PASS" if success else "❌ FAIL"
        print(f"{status} {test_name}: {message}")
        self.test_results[test_name] = {"success": success, "message": message}
        
    def test_api_connectivity(self):
        """Test basic API connectivity"""
        try:
            response = self.session.get(f"{API_BASE}/", timeout=10)
            if response.status_code == 200:
                data = response.json()
                if "message" in data and "Family Quiz API" in data["message"]:
                    self.log_test("API Connectivity", True, f"API responding correctly: {data['message']}")
                    return True
                else:
                    self.log_test("API Connectivity", False, f"Unexpected response: {data}")
                    return False
            else:
                self.log_test("API Connectivity", False, f"HTTP {response.status_code}: {response.text}")
                return False
        except Exception as e:
            self.log_test("API Connectivity", False, f"Connection error: {str(e)}")
            return False
    
    def test_qr_code_generation(self):
        """Test QR code generation endpoint"""
        try:
            response = self.session.get(f"{API_BASE}/qr-code", timeout=10)
            if response.status_code == 200:
                data = response.json()
                required_fields = ["qr_code", "url", "local_ip"]
                
                if all(field in data for field in required_fields):
                    # Verify QR code is base64 encoded image
                    if data["qr_code"].startswith("data:image/png;base64,"):
                        # Verify URL format
                        if "/join" in data["url"]:
                            self.log_test("QR Code Generation", True, f"QR code generated for {data['url']}")
                            return True
                        else:
                            self.log_test("QR Code Generation", False, f"Invalid URL format: {data['url']}")
                            return False
                    else:
                        self.log_test("QR Code Generation", False, "QR code not in base64 format")
                        return False
                else:
                    missing = [f for f in required_fields if f not in data]
                    self.log_test("QR Code Generation", False, f"Missing fields: {missing}")
                    return False
            else:
                self.log_test("QR Code Generation", False, f"HTTP {response.status_code}: {response.text}")
                return False
        except Exception as e:
            self.log_test("QR Code Generation", False, f"Error: {str(e)}")
            return False
    
    def test_template_excel_download(self):
        """Test template Excel file download"""
        try:
            response = self.session.get(f"{API_BASE}/template-excel", timeout=10)
            if response.status_code == 200:
                # Check content type
                content_type = response.headers.get('content-type', '')
                if 'spreadsheet' in content_type or 'excel' in content_type:
                    # Try to parse the Excel file
                    try:
                        wb = openpyxl.load_workbook(BytesIO(response.content))
                        ws = wb.active
                        
                        # Check if it has the expected structure
                        headers = [ws.cell(1, col).value for col in range(1, 9)]
                        expected_headers = ["ID", "Question", "Option A", "Option B", "Option C", "Option D", "Duration (seconds)", "Points"]
                        
                        if headers == expected_headers:
                            # Check if there are sample questions
                            if ws.max_row > 1:
                                self.log_test("Template Excel Download", True, f"Valid Excel template with {ws.max_row-1} sample questions")
                                return True
                            else:
                                self.log_test("Template Excel Download", False, "Template has no sample questions")
                                return False
                        else:
                            self.log_test("Template Excel Download", False, f"Invalid headers: {headers}")
                            return False
                    except Exception as e:
                        self.log_test("Template Excel Download", False, f"Invalid Excel format: {str(e)}")
                        return False
                else:
                    self.log_test("Template Excel Download", False, f"Wrong content type: {content_type}")
                    return False
            else:
                self.log_test("Template Excel Download", False, f"HTTP {response.status_code}: {response.text}")
                return False
        except Exception as e:
            self.log_test("Template Excel Download", False, f"Error: {str(e)}")
            return False
    
    def create_test_excel(self):
        """Create a test Excel file with red cell marking correct answers"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Quiz Questions"
        
        # Headers
        headers = ["ID", "Question", "Option A", "Option B", "Option C", "Option D", "Duration (seconds)", "Points"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Test questions with red marking for correct answers
        from openpyxl.styles import PatternFill
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        
        questions = [
            ["T1", "What is 2+2?", "3", "4", "5", "6", 30, 10],  # Correct: B
            ["T2", "Capital of France?", "London", "Berlin", "Paris", "Madrid", 25, 15],  # Correct: C
        ]
        
        correct_cols = [4, 5]  # B and C columns (0-indexed: 3, 4 -> 1-indexed: 4, 5)
        
        for row, question_data in enumerate(questions, 2):
            for col, value in enumerate(question_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                
                # Mark correct answer in red
                if col == correct_cols[row - 2]:
                    cell.fill = red_fill
        
        # Save to memory
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer
    
    def test_excel_upload_processing(self):
        """Test Excel file upload and processing"""
        try:
            # Create test Excel file
            excel_file = self.create_test_excel()
            
            # Upload the file
            files = {'file': ('test_quiz.xlsx', excel_file.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            response = self.session.post(f"{API_BASE}/upload-excel", files=files, timeout=15)
            
            if response.status_code == 200:
                data = response.json()
                if "message" in data and "questions" in data:
                    questions = data["questions"]
                    if len(questions) == 2:  # We uploaded 2 test questions
                        # Verify question parsing
                        q1 = questions[0]
                        q2 = questions[1]
                        
                        # Check if correct answers were detected from red cells
                        if q1["correct_answer"] == "B" and q2["correct_answer"] == "C":
                            self.log_test("Excel Upload Processing", True, f"Successfully processed {len(questions)} questions with correct red cell detection")
                            return True
                        else:
                            self.log_test("Excel Upload Processing", False, f"Red cell detection failed: Q1={q1['correct_answer']}, Q2={q2['correct_answer']}")
                            return False
                    else:
                        self.log_test("Excel Upload Processing", False, f"Expected 2 questions, got {len(questions)}")
                        return False
                else:
                    self.log_test("Excel Upload Processing", False, f"Invalid response format: {data}")
                    return False
            else:
                self.log_test("Excel Upload Processing", False, f"HTTP {response.status_code}: {response.text}")
                return False
        except Exception as e:
            self.log_test("Excel Upload Processing", False, f"Error: {str(e)}")
            return False
    
    def test_quiz_management_apis(self):
        """Test quiz management endpoints"""
        try:
            # First upload questions (prerequisite)
            excel_file = self.create_test_excel()
            files = {'file': ('test_quiz.xlsx', excel_file.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            upload_response = self.session.post(f"{API_BASE}/upload-excel", files=files, timeout=15)
            
            if upload_response.status_code != 200:
                self.log_test("Quiz Management APIs", False, "Failed to upload questions for testing")
                return False
            
            # Test quiz state endpoint
            state_response = self.session.get(f"{API_BASE}/quiz-state", timeout=10)
            if state_response.status_code != 200:
                self.log_test("Quiz Management APIs", False, f"Quiz state endpoint failed: {state_response.status_code}")
                return False
            
            # Test start quiz
            start_response = self.session.post(f"{API_BASE}/start-quiz", timeout=10)
            if start_response.status_code != 200:
                self.log_test("Quiz Management APIs", False, f"Start quiz failed: {start_response.status_code}")
                return False
            
            # Test pause quiz
            pause_response = self.session.post(f"{API_BASE}/pause-quiz", timeout=10)
            if pause_response.status_code != 200:
                self.log_test("Quiz Management APIs", False, f"Pause quiz failed: {pause_response.status_code}")
                return False
            
            # Test resume quiz
            resume_response = self.session.post(f"{API_BASE}/resume-quiz", timeout=10)
            if resume_response.status_code != 200:
                self.log_test("Quiz Management APIs", False, f"Resume quiz failed: {resume_response.status_code}")
                return False
            
            # Test next question
            next_response = self.session.post(f"{API_BASE}/next-question", timeout=10)
            if next_response.status_code != 200:
                self.log_test("Quiz Management APIs", False, f"Next question failed: {next_response.status_code}")
                return False
            
            # Test scores endpoint
            scores_response = self.session.get(f"{API_BASE}/scores", timeout=10)
            if scores_response.status_code != 200:
                self.log_test("Quiz Management APIs", False, f"Scores endpoint failed: {scores_response.status_code}")
                return False
            
            self.log_test("Quiz Management APIs", True, "All quiz management endpoints working correctly")
            return True
            
        except Exception as e:
            self.log_test("Quiz Management APIs", False, f"Error: {str(e)}")
            return False
    
    async def test_socketio_connection(self):
        """Test Socket.IO connection and real-time messaging"""
        try:
            # Create Socket.IO client with debugging
            sio = socketio.AsyncClient(logger=True, engineio_logger=True)
            
            # Track events
            events_received = {}
            connection_successful = False
            
            @sio.event
            async def connect():
                nonlocal connection_successful
                print("Socket.IO client connected successfully")
                events_received['connect'] = True
                connection_successful = True
            
            @sio.event
            async def connect_error(data):
                print(f"Socket.IO connection error: {data}")
                events_received['connect_error'] = data
            
            @sio.event
            async def disconnect():
                print("Socket.IO client disconnected")
                events_received['disconnect'] = True
            
            @sio.event
            async def player_joined(data):
                print(f"Player joined event received: {data}")
                events_received['player_joined'] = data
            
            @sio.event
            async def questions_loaded(data):
                print(f"Questions loaded event received: {data}")
                events_received['questions_loaded'] = data
            
            @sio.event
            async def quiz_started(data):
                print(f"Quiz started event received: {data}")
                events_received['quiz_started'] = data
            
            print(f"Attempting to connect to Socket.IO server at: {BACKEND_URL}")
            
            # Connect to Socket.IO server with timeout
            try:
                await asyncio.wait_for(sio.connect(BACKEND_URL), timeout=10.0)
            except asyncio.TimeoutError:
                self.log_test("Socket.IO Connection", False, "Connection timeout - Socket.IO server may not be properly configured")
                return False
            
            # Wait for connection confirmation
            await asyncio.sleep(2)
            
            if not connection_successful:
                error_info = events_received.get('connect_error', 'Unknown connection error')
                self.log_test("Socket.IO Connection", False, f"Failed to establish Socket.IO connection: {error_info}")
                return False
            
            self.log_test("Socket.IO Connection", True, "Socket.IO connection established successfully")
            
            # Test player join event
            print("Testing player join event...")
            await sio.emit('join_player', {'name': 'Test Player Alice'})
            
            # Wait for player_joined event
            await asyncio.sleep(3)
            
            if 'player_joined' in events_received:
                player_data = events_received['player_joined']
                if 'player' in player_data and player_data['player']['name'] == 'Test Player Alice':
                    self.log_test("Socket.IO Player Join", True, f"Player join event processed correctly: {player_data['player']['name']}")
                else:
                    self.log_test("Socket.IO Player Join", False, f"Invalid player join response: {player_data}")
            else:
                self.log_test("Socket.IO Player Join", False, "No player_joined event received - Socket.IO events may not be working")
            
            # Test answer submission (this will fail gracefully if no active quiz)
            print("Testing answer submission...")
            await sio.emit('submit_answer', {'answer': 'A'})
            await asyncio.sleep(1)
            
            # Disconnect
            await sio.disconnect()
            
            return connection_successful
            
        except Exception as e:
            error_msg = str(e)
            print(f"Socket.IO test exception: {error_msg}")
            self.log_test("Socket.IO Connection", False, f"Socket.IO connection failed: {error_msg}")
            return False
    
    def run_all_tests(self):
        """Run all backend tests"""
        print("=" * 60)
        print("FAMILY QUIZ BACKEND TESTING")
        print("=" * 60)
        print(f"Testing backend at: {BACKEND_URL}")
        print()
        
        # Test basic connectivity first
        if not self.test_api_connectivity():
            print("\n❌ CRITICAL: API not accessible. Stopping tests.")
            return False
        
        # Run all tests
        tests = [
            self.test_qr_code_generation,
            self.test_template_excel_download,
            self.test_excel_upload_processing,
            self.test_quiz_management_apis,
        ]
        
        for test in tests:
            test()
            time.sleep(0.5)  # Small delay between tests
        
        # Run Socket.IO test separately (async)
        print("\nTesting Socket.IO functionality...")
        try:
            asyncio.run(self.test_socketio_connection())
        except Exception as e:
            self.log_test("Socket.IO Connection", False, f"Socket.IO test failed: {str(e)}")
        
        # Summary
        print("\n" + "=" * 60)
        print("TEST SUMMARY")
        print("=" * 60)
        
        passed = sum(1 for result in self.test_results.values() if result["success"])
        total = len(self.test_results)
        
        print(f"Tests passed: {passed}/{total}")
        
        if passed == total:
            print("🎉 ALL TESTS PASSED!")
            return True
        else:
            print("❌ SOME TESTS FAILED")
            print("\nFailed tests:")
            for test_name, result in self.test_results.items():
                if not result["success"]:
                    print(f"  - {test_name}: {result['message']}")
            return False

def main():
    """Main test execution"""
    tester = BackendTester()
    success = tester.run_all_tests()
    
    if success:
        print("\n✅ Backend testing completed successfully!")
    else:
        print("\n❌ Backend testing completed with failures!")
    
    return success

if __name__ == "__main__":
    main()