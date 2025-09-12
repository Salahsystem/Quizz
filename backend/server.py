from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import socketio
import os
import logging
import asyncio
import json
import socket
import qrcode
from io import BytesIO
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Dict, Optional
import uuid
from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import PatternFill
import base64

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# WebSocket connection manager
class ConnectionManager:
    def __init__(self):
        self.active_connections: Dict[str, WebSocket] = {}
        self.players: Dict[str, Dict] = {}
        self.quiz_state = {
            "status": "waiting",  # waiting, lobby, active, paused, finished
            "current_question": 0,
            "questions": [],
            "quiz_id": None,
            "start_time": None,
            "question_start_time": None
        }

    async def connect(self, websocket: WebSocket, client_id: str):
        await websocket.accept()
        self.active_connections[client_id] = websocket

    def disconnect(self, client_id: str):
        if client_id in self.active_connections:
            del self.active_connections[client_id]
        if client_id in self.players:
            del self.players[client_id]

    async def send_personal_message(self, message: str, client_id: str):
        if client_id in self.active_connections:
            try:
                await self.active_connections[client_id].send_text(message)
            except:
                self.disconnect(client_id)

    async def broadcast(self, message: str):
        disconnected = []
        for client_id, connection in self.active_connections.items():
            try:
                await connection.send_text(message)
            except:
                disconnected.append(client_id)
        
        for client_id in disconnected:
            self.disconnect(client_id)

manager = ConnectionManager()

# Define Models
class QuizQuestion(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    question: str
    option_a: str
    option_b: str
    option_c: str
    option_d: str
    correct_answer: str  # A, B, C, or D
    duration: int
    points: int

class Player(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    score: int = 0
    joined_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class QuizSession(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    questions: List[QuizQuestion]
    players: Dict[str, Player] = {}
    status: str = "waiting"
    current_question: int = 0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PlayerAnswer(BaseModel):
    player_id: str
    answer: str
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Helper function to get local IP
def get_local_ip():
    try:
        # Connect to a remote address to determine local IP
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as s:
            s.connect(("8.8.8.8", 80))
            return s.getsockname()[0]
    except:
        return "localhost"

# Helper function to generate template Excel file
def generate_template_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quiz Questions"
    
    # Headers
    headers = ["ID", "Question", "Option A", "Option B", "Option C", "Option D", "Duration (seconds)", "Points"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Sample questions
    questions = [
        ["Q1", "Quelle est la capitale de la France ?", "Berlin", "Madrid", "Paris", "Rome", 30, 10],
        ["Q2", "Combien font 3 x 4 ?", "10", "12", "14", "16", 20, 5],
        ["Q3", "Qui a peint la Joconde ?", "Picasso", "Léonard de Vinci", "Van Gogh", "Monet", 25, 15]
    ]
    
    # Add sample questions and mark correct answers in red
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    correct_answers = ["C", "B", "B"]  # Paris, 12, Léonard de Vinci
    
    for row, question_data in enumerate(questions, 2):
        for col, value in enumerate(question_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            
            # Mark correct answer in red
            if col >= 3 and col <= 6:  # Options A, B, C, D
                option_letter = chr(ord('A') + (col - 3))
                if option_letter == correct_answers[row - 2]:
                    cell.fill = red_fill
    
    # Save to memory
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer

# Helper function to parse Excel file
def parse_excel_file(file_content: bytes):
    wb = openpyxl.load_workbook(BytesIO(file_content))
    ws = wb.active
    
    questions = []
    
    for row in range(2, ws.max_row + 1):
        try:
            question_id = ws.cell(row=row, column=1).value
            question_text = ws.cell(row=row, column=2).value
            option_a = ws.cell(row=row, column=3).value
            option_b = ws.cell(row=row, column=4).value
            option_c = ws.cell(row=row, column=5).value
            option_d = ws.cell(row=row, column=6).value
            duration = ws.cell(row=row, column=7).value or 30
            points = ws.cell(row=row, column=8).value or 10
            
            if not all([question_text, option_a, option_b, option_c, option_d]):
                continue
            
            # Find correct answer by looking for red cell
            correct_answer = "A"  # default
            for col in range(3, 7):  # Options A, B, C, D
                cell = ws.cell(row=row, column=col)
                if cell.fill.start_color.rgb == "FFFF0000":  # Red color
                    correct_answer = chr(ord('A') + (col - 3))
                    break
            
            questions.append(QuizQuestion(
                id=str(question_id) if question_id else str(uuid.uuid4()),
                question=str(question_text),
                option_a=str(option_a),
                option_b=str(option_b),
                option_c=str(option_c),
                option_d=str(option_d),
                correct_answer=correct_answer,
                duration=int(duration),
                points=int(points)
            ))
        except Exception as e:
            logging.error(f"Error parsing row {row}: {e}")
            continue
    
    return questions

# WebSocket endpoint
@app.websocket("/ws/{client_id}")
async def websocket_endpoint(websocket: WebSocket, client_id: str):
    await manager.connect(websocket, client_id)
    try:
        while True:
            data = await websocket.receive_text()
            message = json.loads(data)
            
            if message["type"] == "join_player":
                player = Player(id=client_id, name=message["name"])
                manager.players[client_id] = player.dict()
                
                await manager.broadcast(json.dumps({
                    "type": "player_joined",
                    "player": player.dict(),
                    "players": list(manager.players.values())
                }))
                
            elif message["type"] == "submit_answer":
                if client_id in manager.players and manager.quiz_state["status"] == "active":
                    player = manager.players[client_id]
                    current_q_idx = manager.quiz_state["current_question"]
                    
                    if current_q_idx < len(manager.quiz_state["questions"]):
                        question = manager.quiz_state["questions"][current_q_idx]
                        is_correct = message["answer"] == question["correct_answer"]
                        
                        if is_correct:
                            player["score"] += question["points"]
                            manager.players[client_id] = player
                        
                        await manager.send_personal_message(json.dumps({
                            "type": "answer_feedback",
                            "correct": is_correct,
                            "correct_answer": question["correct_answer"],
                            "score": player["score"]
                        }), client_id)
            
    except WebSocketDisconnect:
        manager.disconnect(client_id)
        await manager.broadcast(json.dumps({
            "type": "player_left",
            "players": list(manager.players.values())
        }))

# API Routes
@api_router.get("/")
async def root():
    return {"message": "Family Quiz API"}

@api_router.get("/qr-code")
async def get_qr_code():
    local_ip = get_local_ip()
    frontend_url = f"http://{local_ip}:3000/join"
    
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(frontend_url)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    img_buffer = BytesIO()
    img.save(img_buffer, format='PNG')
    img_buffer.seek(0)
    
    # Convert to base64 for JSON response
    img_base64 = base64.b64encode(img_buffer.getvalue()).decode()
    
    return {
        "qr_code": f"data:image/png;base64,{img_base64}",
        "url": frontend_url,
        "local_ip": local_ip
    }

@api_router.get("/template-excel")
async def download_template():
    excel_buffer = generate_template_excel()
    
    return StreamingResponse(
        BytesIO(excel_buffer.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=quiz_template.xlsx"}
    )

@api_router.post("/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files are allowed")
    
    content = await file.read()
    questions = parse_excel_file(content)
    
    if not questions:
        raise HTTPException(status_code=400, detail="No valid questions found in Excel file")
    
    # Store questions in quiz state
    manager.quiz_state["questions"] = [q.dict() for q in questions]
    manager.quiz_state["quiz_id"] = str(uuid.uuid4())
    
    await manager.broadcast(json.dumps({
        "type": "questions_loaded",
        "count": len(questions),
        "questions": [q.dict() for q in questions]
    }))
    
    return {"message": f"Successfully loaded {len(questions)} questions", "questions": questions}

@api_router.post("/start-quiz")
async def start_quiz():
    if not manager.quiz_state["questions"]:
        raise HTTPException(status_code=400, detail="No questions loaded")
    
    manager.quiz_state["status"] = "active"
    manager.quiz_state["current_question"] = 0
    manager.quiz_state["start_time"] = datetime.now(timezone.utc)
    
    await manager.broadcast(json.dumps({
        "type": "quiz_started",
        "status": "active"
    }))
    
    # Send first question
    await send_current_question()
    
    return {"message": "Quiz started"}

@api_router.post("/next-question")
async def next_question():
    if manager.quiz_state["status"] != "active":
        raise HTTPException(status_code=400, detail="Quiz is not active")
    
    manager.quiz_state["current_question"] += 1
    
    if manager.quiz_state["current_question"] >= len(manager.quiz_state["questions"]):
        # Quiz finished
        manager.quiz_state["status"] = "finished"
        await manager.broadcast(json.dumps({
            "type": "quiz_finished",
            "final_scores": list(manager.players.values())
        }))
        return {"message": "Quiz finished"}
    
    await send_current_question()
    return {"message": "Next question sent"}

@api_router.post("/pause-quiz")
async def pause_quiz():
    manager.quiz_state["status"] = "paused"
    await manager.broadcast(json.dumps({
        "type": "quiz_paused"
    }))
    return {"message": "Quiz paused"}

@api_router.post("/resume-quiz")
async def resume_quiz():
    manager.quiz_state["status"] = "active"
    await manager.broadcast(json.dumps({
        "type": "quiz_resumed"
    }))
    return {"message": "Quiz resumed"}

@api_router.get("/quiz-state")
async def get_quiz_state():
    return {
        "status": manager.quiz_state["status"],
        "current_question": manager.quiz_state["current_question"],
        "total_questions": len(manager.quiz_state["questions"]),
        "players": list(manager.players.values())
    }

@api_router.get("/scores")
async def get_scores():
    sorted_players = sorted(manager.players.values(), key=lambda x: x["score"], reverse=True)
    return {"scores": sorted_players}

async def send_current_question():
    if manager.quiz_state["current_question"] < len(manager.quiz_state["questions"]):
        question = manager.quiz_state["questions"][manager.quiz_state["current_question"]]
        manager.quiz_state["question_start_time"] = datetime.now(timezone.utc)
        
        await manager.broadcast(json.dumps({
            "type": "question",
            "question": {
                "id": question["id"],
                "question": question["question"],
                "option_a": question["option_a"],
                "option_b": question["option_b"],
                "option_c": question["option_c"],
                "option_d": question["option_d"],
                "duration": question["duration"],
                "points": question["points"]
            },
            "question_number": manager.quiz_state["current_question"] + 1,
            "total_questions": len(manager.quiz_state["questions"])
        }))

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()